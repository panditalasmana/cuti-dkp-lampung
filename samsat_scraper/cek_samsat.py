# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

"""
=======================================================
  SAMSAT PKB DKI Jakarta - Cek Otomatis v3
  Fitur: CAPTCHA Otomatis (CapSolver) + Resume + Progress
=======================================================
  SETUP:
  1. Daftar di https://capsolver.com → top up saldo minimal $5
  2. Copy API Key Anda → isi di variabel CAPSOLVER_API_KEY di bawah
  3. Isi file 'daftar_nopol.xlsx' kolom NOPOL
  4. Jalankan: python cek_samsat.py
  5. Program 100% otomatis - bisa ditinggal!
=======================================================
"""

import time
import re
import os
import json
import glob
import requests
from datetime import datetime, timedelta

# ─────────────────────────────────────────────
#  CEK & INSTALL LIBRARY
# ─────────────────────────────────────────────
def install_requirements():
    import subprocess
    packages = {
        "selenium"         : "selenium",
        "webdriver_manager": "webdriver-manager",
        "openpyxl"         : "openpyxl",
        "requests"         : "requests",
    }
    for import_name, pkg_name in packages.items():
        try:
            __import__(import_name)
        except ImportError:
            print(f"[INSTALL] Menginstall {pkg_name}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg_name, "-q"])

install_requirements()

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════
#  KONFIGURASI UTAMA — UBAH DI SINI
# ════════════════════════════════════════════════
CAPSOLVER_API_KEY  = "ISI_API_KEY_ANDA_DISINI"   # <── Ganti dengan API Key CapSolver Anda
                                                   # Daftar gratis di: https://capsolver.com

INPUT_FILE         = "daftar_nopol.xlsx"
NOPOL_COLUMN       = "NOPOL"
URL                = "https://samsat-pkb2.jakarta.go.id/"
RECAPTCHA_SITE_KEY = "6Le5S-ojAAAAALzLdTc2YEeGRbT43dVRl58GA7dR"  # Site key Samsat Jakarta
OUTPUT_PREFIX      = "hasil_samsat"
RESULT_TIMEOUT     = 15    # Detik tunggu halaman hasil

# ════════════════════════════════════════════════
#  CAPSOLVER — PECAHKAN reCAPTCHA OTOMATIS
# ════════════════════════════════════════════════
def cek_saldo_capsolver():
    """Cek saldo CapSolver sebelum mulai."""
    try:
        resp = requests.post("https://api.capsolver.com/getBalance", json={
            "clientKey": CAPSOLVER_API_KEY
        }, timeout=10)
        data = resp.json()
        if data.get("errorId") == 0:
            saldo = data.get("balance", 0)
            print(f"[CapSolver] Saldo: ${saldo:.4f}")
            if float(saldo) < 0.01:
                print("[PERINGATAN] Saldo CapSolver hampir habis! Top up di https://capsolver.com")
            return True
        else:
            print(f"[CapSolver] ERROR: {data.get('errorDescription')}")
            return False
    except Exception as e:
        print(f"[CapSolver] Gagal cek saldo: {e}")
        return False

def solve_recaptcha_capsolver(url, site_key, max_tunggu=120):
    """
    Kirim gambar CAPTCHA ke CapSolver, tunggu hasilnya.
    Return: token string jika berhasil, None jika gagal.
    """
    print("   [CapSolver] Mengirim permintaan solve reCAPTCHA...")
    try:
        # Buat task
        resp = requests.post("https://api.capsolver.com/createTask", json={
            "clientKey": CAPSOLVER_API_KEY,
            "task": {
                "type"           : "ReCaptchaV2TaskProxyless",
                "websiteURL"     : url,
                "websiteKey"     : site_key,
                "isInvisible"    : False,
            }
        }, timeout=30)

        data = resp.json()
        if data.get("errorId") != 0:
            print(f"   [CapSolver] Gagal buat task: {data.get('errorDescription')}")
            return None

        task_id = data.get("taskId")
        print(f"   [CapSolver] Task ID: {task_id} | Menunggu hasil...")

        # Poll hasil
        start = time.time()
        while time.time() - start < max_tunggu:
            time.sleep(4)  # Tunggu 4 detik setiap polling
            resp2 = requests.post("https://api.capsolver.com/getTaskResult", json={
                "clientKey": CAPSOLVER_API_KEY,
                "taskId"   : task_id
            }, timeout=15)

            result = resp2.json()
            status = result.get("status")

            if status == "ready":
                token = result["solution"]["gRecaptchaResponse"]
                print(f"   [CapSolver] Token berhasil didapat!")
                return token
            elif status == "processing":
                elapsed = int(time.time() - start)
                print(f"   [CapSolver] Sedang diproses... ({elapsed}s)")
            else:
                print(f"   [CapSolver] Status tidak dikenal: {status}")
                return None

        print("   [CapSolver] Timeout - tidak dapat token dalam batas waktu.")
        return None

    except requests.exceptions.RequestException as e:
        print(f"   [CapSolver] Error koneksi: {e}")
        return None
    except Exception as e:
        print(f"   [CapSolver] Error: {e}")
        return None

def inject_token_ke_browser(driver, token):
    """Inject token reCAPTCHA ke dalam browser via JavaScript."""
    try:
        # Set response token di textarea tersembunyi reCAPTCHA
        driver.execute_script(f"""
            document.getElementById('g-recaptcha-response').innerHTML = '{token}';
        """)
        # Trigger callback verifyRecaptchaCallback
        driver.execute_script(f"""
            window.verifyRecaptchaCallback('{token}');
        """)
        # Set juga pada input[data-recaptcha]
        driver.execute_script(f"""
            var inputs = document.querySelectorAll('input[data-recaptcha]');
            inputs.forEach(function(inp) {{
                inp.value = '{token}';
                inp.dispatchEvent(new Event('change'));
            }});
        """)
        print("   [Token] Token berhasil di-inject ke browser.")
        return True
    except Exception as e:
        print(f"   [Token] Gagal inject token: {e}")
        return False

# ─────────────────────────────────────────────
#  PARSE NOPOL: B5834BWN → angka=5834, huruf=BWN
# ─────────────────────────────────────────────
def parse_nopol(nopol_raw):
    nopol = nopol_raw.strip().upper().replace(" ", "")
    match = re.match(r'^([A-Z]+)(\d{1,4})([A-Z]*)$', nopol)
    if match:
        return match.group(2), match.group(3)
    return None, None

# ─────────────────────────────────────────────
#  PROGRESS BAR
# ─────────────────────────────────────────────
def tampil_progress(idx, total, start_time, nopol):
    persen = idx / total * 100
    sisa   = total - idx
    elapsed = time.time() - start_time
    if idx > 1:
        per_item   = elapsed / (idx - 1)
        sisa_str   = str(timedelta(seconds=int(sisa * per_item)))
    else:
        sisa_str = "menghitung..."

    bar_len = 30
    filled  = int(bar_len * idx // total)
    bar     = "#" * filled + "-" * (bar_len - filled)

    print(f"\n{'=' * 60}")
    print(f"  [{bar}] {persen:.1f}%")
    print(f"  Selesai: {idx-1}/{total}  |  Sisa: {sisa} NOPOL")
    print(f"  Estimasi sisa waktu: {sisa_str}")
    print(f"  Sedang proses: {nopol}")
    print(f"{'=' * 60}")

# ─────────────────────────────────────────────
#  FILE OUTPUT (RESUME)
# ─────────────────────────────────────────────
HEADERS = [
    "No", "NOPOL Input", "Nopol", "Nama", "Alamat",
    "No. Rangka", "No. Mesin", "Merek / Type",
    "Warna Kendaraan", "Bhn Bakar / Cylinder",
    "Nilai Jual", "Masa Berlaku STNK",
    "Jatuh Tempo Pajak", "PKB Pokok", "PKB Denda",
    "SWDKLLJ", "SWDKLLJ Denda", "TOTAL", "STATUS",
    "Kendaraan ke", "No BPKB", "Model / Pembuatan",
    "Warna TNKB", "Waktu Cek", "Keterangan"
]

def cari_atau_buat_output():
    files = sorted(glob.glob(f"{OUTPUT_PREFIX}_*.xlsx"))
    if files:
        latest = files[-1]
        print(f"\n[RESUME] File ditemukan: {latest}")
        print(f"         Program akan MELANJUTKAN data yang belum diproses.")
        return latest, False
    timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f"{OUTPUT_PREFIX}_{timestamp}.xlsx"
    return output_file, True

def baca_sudah_diproses(output_file):
    sudah = set()
    if not os.path.exists(output_file):
        return sudah
    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        header_row = [str(c.value).strip() if c.value else "" for c in ws[1]]
        if "NOPOL Input" not in header_row:
            return sudah
        col_idx = header_row.index("NOPOL Input")
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_idx]
            if val:
                sudah.add(str(val).strip().upper())
        wb.close()
    except Exception as e:
        print(f"[WARN] Gagal baca file output: {e}")
    return sudah

def buat_excel_baru(output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hasil Samsat"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    ws.row_dimensions[1].height = 30
    wb.save(output_file)
    print(f"[OK] File output baru: {output_file}")

def tulis_baris(output_file, row_data, excel_row_num):
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    err_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    thin     = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    row_values = [row_data.get(h, "") for h in HEADERS]
    ws.append(row_values)
    last_row   = ws.max_row
    status     = str(row_data.get("STATUS", "")).upper()
    is_error   = any(k in status for k in ["ERROR", "TIDAK DITEMUKAN", "TIMEOUT", "GAGAL", "FORMAT"])
    for col_idx in range(1, len(HEADERS) + 1):
        cell           = ws.cell(row=last_row, column=col_idx)
        cell.font      = Font(size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = thin
        cell.fill      = err_fill if is_error else (alt_fill if excel_row_num % 2 == 0 else PatternFill())
    wb.save(output_file)
    wb.close()

# ─────────────────────────────────────────────
#  BROWSER
# ─────────────────────────────────────────────
def buat_browser():
    print("[Browser] Menyiapkan Chrome...")
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    return driver

# ─────────────────────────────────────────────
#  SCRAPE DATA
# ─────────────────────────────────────────────
LABEL_MAP = {
    "nopol"                : "Nopol",
    "nama"                 : "Nama",
    "alamat"               : "Alamat",
    "no. rangka / mesin"   : "No. Rangka / Mesin",
    "no. rangka"           : "No. Rangka",
    "no. mesin"            : "No. Mesin",
    "merek / type"         : "Merek / Type",
    "warna kendaraan"      : "Warna Kendaraan",
    "bhn bakar / cylinder" : "Bhn Bakar / Cylinder",
    "nilai jual"           : "Nilai Jual",
    "masa berlaku stnk"    : "Masa Berlaku STNK",
    "jatuh tempo pajak"    : "Jatuh Tempo Pajak",
    "pkb pokok"            : "PKB Pokok",
    "pkb denda"            : "PKB Denda",
    "swdkllj denda"        : "SWDKLLJ Denda",
    "swdkllj"              : "SWDKLLJ",
    "total"                : "TOTAL",
    "status"               : "STATUS",
    "kendaraan ke"         : "Kendaraan ke",
    "no bpkb"              : "No BPKB",
    "model / pembuatan"    : "Model / Pembuatan",
    "warna tnkb"           : "Warna TNKB",
}

def scrape_hasil(driver):
    hasil = {}
    try:
        WebDriverWait(driver, RESULT_TIMEOUT).until(
            lambda d: (
                len(d.find_elements(By.TAG_NAME, "td")) > 4
                or "tidak ditemukan" in d.page_source.lower()
            )
        )
        time.sleep(1)
        tds = driver.find_elements(By.TAG_NAME, "td")
        i = 0
        while i < len(tds) - 1:
            label_raw = tds[i].text.strip().lower().rstrip(':').rstrip('*)')
            nilai     = tds[i + 1].text.strip()
            matched   = None
            for k, v in LABEL_MAP.items():
                if k == label_raw or k in label_raw or label_raw in k:
                    matched = v
                    break
            if matched:
                if matched == "No. Rangka / Mesin" and "/" in nilai:
                    parts = nilai.split("/", 1)
                    hasil["No. Rangka"] = parts[0].strip()
                    hasil["No. Mesin"]  = parts[1].strip()
                elif matched not in hasil:
                    hasil[matched] = nilai
            i += 1
        if not hasil:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            if "tidak ditemukan" in body or "data not found" in body:
                hasil["STATUS"]     = "TIDAK DITEMUKAN"
                hasil["Keterangan"] = "Data kendaraan tidak ditemukan"
            else:
                hasil["STATUS"]     = "GAGAL SCRAPE"
                hasil["Keterangan"] = "Halaman hasil tidak terbaca"
    except TimeoutException:
        hasil["STATUS"]     = "TIMEOUT HASIL"
        hasil["Keterangan"] = "Halaman hasil tidak muncul"
    except Exception as e:
        hasil["STATUS"]     = "ERROR"
        hasil["Keterangan"] = str(e)[:200]
    return hasil

# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  SAMSAT PKB DKI Jakarta - Cek Otomatis v3")
    print("  Mode: CAPTCHA Otomatis via CapSolver")
    print("=" * 60)

    # Validasi API Key
    if CAPSOLVER_API_KEY == "ISI_API_KEY_ANDA_DISINI":
        print("\n[ERROR] API Key CapSolver belum diisi!")
        print("  1. Daftar di https://capsolver.com")
        print("  2. Top up saldo minimal $5 (~Rp 80.000)")
        print("  3. Buka file cek_samsat.py")
        print("  4. Ganti 'ISI_API_KEY_ANDA_DISINI' dengan API Key Anda")
        input("\nTekan Enter untuk keluar...")
        return

    # Cek saldo
    if not cek_saldo_capsolver():
        print("\n[ERROR] Tidak bisa terhubung ke CapSolver. Cek API Key dan koneksi internet.")
        input("\nTekan Enter untuk keluar...")
        return

    # Baca input Excel
    if not os.path.exists(INPUT_FILE):
        print(f"\n[ERROR] File '{INPUT_FILE}' tidak ditemukan!")
        print(f"   Jalankan dulu: python buat_contoh_excel.py")
        input("\nTekan Enter untuk keluar...")
        return

    print(f"\n[INFO] Membaca: {INPUT_FILE}")
    wb_in   = openpyxl.load_workbook(INPUT_FILE)
    ws_in   = wb_in.active
    headers = [str(c.value).strip() if c.value else "" for c in ws_in[1]]
    if NOPOL_COLUMN not in headers:
        print(f"[ERROR] Kolom '{NOPOL_COLUMN}' tidak ditemukan! Kolom ada: {[h for h in headers if h]}")
        input("\nTekan Enter untuk keluar...")
        return

    nopol_idx   = headers.index(NOPOL_COLUMN)
    semua_nopol = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        val = row[nopol_idx]
        if val and str(val).strip():
            semua_nopol.append(str(val).strip().upper())
    wb_in.close()

    total = len(semua_nopol)
    est_detik = total * 30  # estimasi 30 detik per NOPOL (termasuk API call)
    print(f"[INFO] Total NOPOL: {total}")
    print(f"[INFO] Estimasi waktu: {str(timedelta(seconds=est_detik))} (otomatis, bisa ditinggal!)")

    # Resume
    output_file, is_baru = cari_atau_buat_output()
    if is_baru:
        buat_excel_baru(output_file)
    sudah = baca_sudah_diproses(output_file)
    if sudah:
        print(f"[RESUME] {len(sudah)} NOPOL sudah diproses, dilewati.")

    antrian = [n for n in semua_nopol if n not in sudah]
    print(f"[INFO] Antrian sisa: {len(antrian)} NOPOL\n")

    if not antrian:
        print("[OK] Semua NOPOL sudah diproses!")
        print(f"     Hasil: {output_file}")
        input("\nTekan Enter untuk keluar...")
        return

    print("Tekan Enter untuk mulai (program berjalan otomatis)...")
    input()

    driver     = buat_browser()
    berhasil   = len(sudah)
    gagal      = 0
    excel_row  = len(sudah) + 1
    start_time = time.time()

    for idx_q, nopol_raw in enumerate(antrian, 1):
        excel_row  += 1
        idx_global  = len(sudah) + idx_q

        tampil_progress(idx_global, total, start_time, nopol_raw)

        angka, huruf = parse_nopol(nopol_raw)
        if not angka:
            print(f"   [SKIP] Format tidak valid: '{nopol_raw}'")
            tulis_baris(output_file, {
                "No": idx_global, "NOPOL Input": nopol_raw,
                "STATUS": "FORMAT TIDAK VALID",
                "Keterangan": "Contoh format benar: B5834BWN",
                "Waktu Cek": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }, excel_row)
            gagal += 1
            continue

        try:
            driver.get(URL)
            time.sleep(2)

            # Isi form
            driver.find_element(By.NAME, "nopa").clear()
            driver.find_element(By.NAME, "nopa").send_keys(angka)
            time.sleep(0.3)
            driver.find_element(By.NAME, "noph").clear()
            driver.find_element(By.NAME, "noph").send_keys(huruf)
            print(f"   [Form] Angka={angka}, Huruf={huruf}")

            # Solve CAPTCHA otomatis
            token = solve_recaptcha_capsolver(URL, RECAPTCHA_SITE_KEY)

            if not token:
                tulis_baris(output_file, {
                    "No": idx_global, "NOPOL Input": nopol_raw,
                    "STATUS": "CAPTCHA GAGAL",
                    "Keterangan": "CapSolver tidak bisa memecahkan CAPTCHA",
                    "Waktu Cek": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                }, excel_row)
                gagal += 1
                continue

            # Inject token ke browser
            injected = inject_token_ke_browser(driver, token)
            if not injected:
                tulis_baris(output_file, {
                    "No": idx_global, "NOPOL Input": nopol_raw,
                    "STATUS": "INJECT TOKEN GAGAL",
                    "Keterangan": "Gagal inject token CAPTCHA ke browser",
                    "Waktu Cek": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                }, excel_row)
                gagal += 1
                continue

            time.sleep(1)

            # Klik tombol Cari
            driver.find_element(By.ID, "submit").click()
            print("   [Cari] Menunggu hasil...")
            time.sleep(2)

            # Scrape data
            hasil = scrape_hasil(driver)
            hasil["No"]          = idx_global
            hasil["NOPOL Input"] = nopol_raw
            hasil["Waktu Cek"]   = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if "Keterangan" not in hasil:
                hasil["Keterangan"] = "Berhasil"

            print(f"   [Data] Nama   : {hasil.get('Nama', '-')}")
            print(f"   [Data] Rangka : {hasil.get('No. Rangka', '-')}")
            print(f"   [Data] Mesin  : {hasil.get('No. Mesin', '-')}")
            print(f"   [Data] Status : {hasil.get('STATUS', '-')}")
            print(f"   [Data] Total  : {hasil.get('TOTAL', '-')}")

            tulis_baris(output_file, hasil, excel_row)
            berhasil += 1

        except KeyboardInterrupt:
            print("\n\n[STOP] Program dihentikan.")
            print(f"       Data tersimpan di: {output_file}")
            print(f"       Jalankan ulang untuk melanjutkan.")
            driver.quit()
            input("\nTekan Enter untuk keluar...")
            return
        except Exception as e:
            print(f"   [ERROR] {e}")
            tulis_baris(output_file, {
                "No": idx_global, "NOPOL Input": nopol_raw,
                "STATUS": "ERROR", "Keterangan": str(e)[:200],
                "Waktu Cek": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }, excel_row)
            gagal += 1

        if idx_q < len(antrian):
            time.sleep(2)

    driver.quit()

    # Rapikan kolom
    try:
        wb_f = openpyxl.load_workbook(output_file)
        ws_f = wb_f.active
        for col in ws_f.columns:
            max_w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws_f.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 4, 40)
        wb_f.save(output_file)
        wb_f.close()
    except:
        pass

    print(f"\n{'=' * 60}")
    print(f"  SELESAI SEMUA!")
    print(f"  Berhasil : {berhasil} NOPOL")
    print(f"  Gagal    : {gagal} NOPOL")
    print(f"  File hasil: {output_file}")
    print(f"{'=' * 60}")
    input("\nTekan Enter untuk keluar...")

if __name__ == "__main__":
    main()
