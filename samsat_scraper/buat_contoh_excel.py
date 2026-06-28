# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
"""
Skrip kecil untuk membuat file Excel contoh 'daftar_nopol.xlsx'
Jalankan sekali saja: python buat_contoh_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Daftar NOPOL"

# Header
ws["A1"] = "NOPOL"
ws["A1"].font  = Font(bold=True, color="FFFFFF")
ws["A1"].fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center")

# Contoh data
contoh = [
    "B5834BWN",
    "B1234ABC",
    "B9876XYZ",
]
for i, nopol in enumerate(contoh, 2):
    ws[f"A{i}"] = nopol

ws.column_dimensions["A"].width = 20
wb.save("daftar_nopol.xlsx")
print("[OK] File 'daftar_nopol.xlsx' berhasil dibuat!")
print("   Buka file tersebut, isi kolom NOPOL dengan daftar plat Anda,")
print("   lalu jalankan 'python cek_samsat.py'")
